import itertools
import os

import numpy as np
from keras.layers import Dense
from keras.models import Sequential
from openpyxl import load_workbook

os.environ['TF_CPP_MIN_LOG_LEVEL'] = '2'


# класс для хранения данных из таблицы
class Data:
    P205: list[float]
    K20: list[float]
    hydrolytic_acid: list[float]
    PH_water: list[float]
    PH_salt: list[float]
    humus: list[float]

    red: list[float]
    green: list[float]
    blue: list[float]

    ch_8: list[float]
    ch_7: list[float]
    ch_6: list[float]
    ch_5: list[float]
    ch_4: list[float]
    ch_3: list[float]
    ch_2: list[float]
    ch_1: list[float]


# Загрузка таблицы в обьект data
wb = load_workbook("./i_kanalov_not_norm.xlsx")
sheet = wb[wb.sheetnames[0]]
count = sum(1 for _ in itertools.takewhile(lambda y: sheet[f"A{y}"].value is not None, itertools.count(2)))
data = Data()
data.P205 = [sheet[f"C{x}"].value for x in range(2, 2 + count)]
data.K20 = [sheet[f"D{x}"].value for x in range(2, 2 + count)]
data.hydrolytic_acid = [sheet[f"E{x}"].value for x in range(2, 2 + count)]
data.PH_water = [sheet[f"F{x}"].value for x in range(2, 2 + count)]
data.humus = [sheet[f"G{x}"].value for x in range(2, 2 + count)]
data.PH_salt = [sheet[f"H{x}"].value for x in range(2, 2 + count)]
data.red = [sheet[f"J{x}"].value for x in range(2, 2 + count)]
data.green = [sheet[f"K{x}"].value for x in range(2, 2 + count)]
data.blue = [sheet[f"L{x}"].value for x in range(2, 2 + count)]
data.ch_8 = [sheet[f"M{x}"].value for x in range(2, 2 + count)]
data.ch_7 = [sheet[f"N{x}"].value for x in range(2, 2 + count)]
data.ch_6 = [sheet[f"O{x}"].value for x in range(2, 2 + count)]
data.ch_5 = [sheet[f"P{x}"].value for x in range(2, 2 + count)]
data.ch_4 = [sheet[f"Q{x}"].value for x in range(2, 2 + count)]
data.ch_3 = [sheet[f"R{x}"].value for x in range(2, 2 + count)]
data.ch_2 = [sheet[f"S{x}"].value for x in range(2, 2 + count)]
data.ch_1 = [sheet[f"T{x}"].value for x in range(2, 2 + count)]

InputArr = []
OutputArr = []
# for i in range(count):
#     InputArr.append([data.P205[i]/1000, data.K20[i]/1000, data.PH_salt[i]/10, data.PH_water[i]/10, data.hydrolytic_acid[i]/100,
#                      data.red[i]/256, data.green[i]/256, data.blue[i]/256, data.ch_8[i]/5000, data.ch_7[i]/5000, data.ch_6[i]/5000,
#                      data.ch_5[i]/5000, data.ch_4[i]/5000, data.ch_3[i]/5000, data.ch_2[i]/5000, data.ch_1[i]/5000])
#     OutputArr.append([data.humus[i]/100])

# заполнение массивов с данными выборки - InputArray и гумуса - OutputArray, при этом данные нормализуются
for i in range(count):
    InputArr.append([data.P205[i] / 1000, data.K20[i] / 1000, data.PH_salt[i] / 10, data.PH_water[i] / 10,
                     data.hydrolytic_acid[i] / 10,
                     data.red[i] / 256, data.green[i] / 256, data.blue[i] / 256, data.ch_8[i] / 5000,
                     data.ch_7[i] / 5000, data.ch_6[i] / 5000,
                     data.ch_5[i] / 5000, data.ch_4[i] / 5000, data.ch_3[i] / 5000, data.ch_2[i] / 5000,
                     data.ch_1[i] / 5000])
    OutputArr.append([data.humus[i] / 100])

InputArr = np.array(InputArr, dtype=float)
OutputArr = np.array(OutputArr, dtype=float)

# создание слоев нейронной сети
model = Sequential()
model.add(Dense(count * 2, input_dim=16, activation='relu'))
model.add(Dense(count * 2, activation='relu'))
model.add(Dense(count * 2, activation='relu'))
model.add(Dense(count * 2, activation='relu'))
model.add(Dense(1, activation='sigmoid'))

# окончательное создание сети с выбором ф-ии ошибки, заполнением данных обучения и выбором количества итераций
model.compile(loss='MSE', optimizer='adam', metrics=['accuracy'])
model.fit(InputArr[:67], OutputArr[:67], epochs=250)
loss, accuracy = model.evaluate(InputArr[:67], OutputArr[:67])
print(f"Точность модели: {accuracy * 100:.2f}%")
predictions = model.predict(InputArr)  # предсказываем
print(abs(predictions - OutputArr) / OutputArr * 100)  # в конце выводим относительную погрешность предсказания
model.save('16_4layer-2_3-_MSE_25000')  # сохранение модели

# model_loaded = keras.models.load_model('16_4layer-2_3-_MSE_25000')
# predictions = model_loaded.predict(InputArr)
# print(abs(predictions-OutputArr)/OutputArr*100)
